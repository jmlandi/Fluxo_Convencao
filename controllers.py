from flask import render_template, redirect, request, flash
from models import Cadastro, db
from datetime import datetime
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pytz

class Controller():

  ############ INDEX ###########
  def index():
    return render_template('index.html')

  def saida():
    return render_template('saida.html')

  ############ CREATE ###########
  def create():

    # Ativando Planilha
    wb = Workbook()
    wb = load_workbook('cadastros.xlsx')
    ws = wb.active
    
    # Salvando dados da planilha -> Dicionario -> Lista
    lista = []
    for row in ws.values:
      cont = 0
      dic = {}
      for value in row:
        if cont == 0:
          dic['Codigo'] = value
        if cont == 1:
          dic['Nome'] = value
        cont += 1
      lista.append(dic)


    #Pegando dados do index
    fluxo = request.form.get('fluxo')
    number_input = request.form.get('number')
    number = str(number_input)
    
      
    # Exportando dados
    for row in lista:
      codigo_ws = row['Codigo']
      codigo = str(codigo_ws)
      nome = row['Nome']
      sptz = pytz.timezone("America/Sao_Paulo") 
      dt = datetime.now(sptz)
      data = dt.strftime("%d/%m/%Y %H:%M:%S")
    
      if codigo == number: 
        novo_cadastro = Cadastro(name=nome,number=codigo,datetime=data, fluxo=fluxo)
        db.session.add(novo_cadastro)
        db.session.commit()
        flash(f'{nome}')
        if fluxo == 'saida':
          return redirect('/saida#digite')
        return redirect('/#digite')
        
    flash('Código não cadastrado.')
    return redirect('/#digite')

  def cadastros():
    cadastros = Cadastro.query.all()
    return render_template('cadastrados.html', cadastros=cadastros)

  def delete(id):
    cadastro = Cadastro.query.filter_by(id=id).first()
    db.session.delete(cadastro)
    db.session.commit()
    return redirect('/cadastros')
    
